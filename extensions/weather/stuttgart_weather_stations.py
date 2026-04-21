import datetime
import os
from pathlib import Path
import shutil
import openpyxl
import pyarrow as pa
import pyarrow.parquet as pq
import pyarrow_ops as pa_ops
import requests

import pytz

from utils.data_pipeline_logger import DataPipelineLogger


class StuttgartWeatherStations:
    """Extension for Stuttgart weather stations."""

    def __init__(
        self,
        extension_data_dir_path,
        meta_data_dir_path,
        input_data_dir_path,
        logs_data_dir_path,
        start_date,
        end_date,
        locations,
    ):
        self.extension_data_dir_path = extension_data_dir_path + "/weather"
        self.meta_data_dir_path = meta_data_dir_path
        self.log_file = Path(logs_data_dir_path) / "logs.log"
        self.extension_temp_dir_path = extension_data_dir_path + "/tmp"

        # Setup logger
        self.logger = DataPipelineLogger.get_logger(
            name=self.__class__.__name__, log_file_path=self.log_file
        )

        os.makedirs(self.extension_data_dir_path, exist_ok=True)
        os.makedirs(self.extension_temp_dir_path, exist_ok=True)

        # start and end date are in format YYYYMMDD. they need to be converted to yyyy-mm-dd for the Stuttgart api
        self.start_date = start_date
        self.end_date = end_date
        self.download_station_ids = [
            "S-Mitte-AfU-Halbstd.-Werte",
            "S-Bad-C-Branddirektion-Halbstd.-Werte",
            "S-Muehlhausen-HKW",
            "S-Vaihingen-UNI",
        ]

        self.station_ids = [
            "S-Mitte_AfU_Halbstd.-Werte",  # Stuttgart Mitte https://www.stadtklima-stuttgart.de/index.php?klima_messdaten_download_afu, (48.774, 9.186)
            "S-Bad-Cannstatt_Branddirektion_Halbstd.-Werte",  # Stuttgart Bad Cannstatt https://www.stadtklima-stuttgart.de/index.php?klima_messdaten_download_fw (48.797, 9.219)
            "S-Mue-HKW",  # Stuttgart Mühlhausen https://www.stadtklima-stuttgart.de/index.php?klima_messdaten_download_hkw, (48.847, 9.241)
            "S-Vai-Uni",  # Stuttgart Vaihingen https://www.stadtklima-stuttgart.de/index.php?klima_messdaten_download_vaih (48.747, 9.107)
        ]

    @property
    def start_timestamp(self):
        """Convert start_date to timestamp once and cache it"""
        if not hasattr(self, "_start_timestamp"):
            self._start_timestamp = int(
                datetime.datetime.strptime(self.start_date, "%Y%m%d")
                .replace(tzinfo=pytz.UTC)
                .timestamp()
            )
        return self._start_timestamp

    @property
    def end_timestamp(self):
        """Convert end_date to timestamp once and cache it"""
        if not hasattr(self, "_end_timestamp"):
            self._end_timestamp = int(
                datetime.datetime.strptime(self.end_date, "%Y%m%d")
                .replace(tzinfo=pytz.UTC)
                .timestamp()
            )
        return self._end_timestamp

    @property
    def start_datetime(self):
        """Convert start_date to datetime once and cache it"""
        if not hasattr(self, "_start_datetime"):
            self._start_datetime = datetime.datetime.strptime(self.start_date, "%Y%m%d")
        return self._start_datetime

    @property
    def end_datetime(self):
        """Convert end_date to datetime once and cache it"""
        if not hasattr(self, "_end_datetime"):
            self._end_datetime = datetime.datetime.strptime(self.end_date, "%Y%m%d")
        return self._end_datetime

    def run(self):
        """Main method to run the Stuttgart weather data processing."""

        self.logger.info(
            "Starting weather data processing for Stuttgart specific stations"
        )
        # Step 1: Download and process weather data
        self.download_weather_data()

        # Step 2: Process the downloaded weather data
        self.process_weather_data()

        # Step 3: Export weather data to parquet format
        self.export_weather_data_to_parquet()

        self.logger.info(
            "Completed weather data processing for Stuttgart specific stations"
        )

    def download_weather_data(self):
        """Download weather data for Stuttgart stations."""
        self.logger.info("Downloading weather data for Stuttgart stations...")
        # "https://www.stadtklima-stuttgart.de/stadtklima_filestorage/download/luft/S-Mitte-AfU-Monatswerte-2019.zip
        base_url = (
            "https://www.stadtklima-stuttgart.de/stadtklima_filestorage/download/luft/"
        )

        years = range(self.start_datetime.year, self.end_datetime.year + 1)

        for weather_station in self.download_station_ids:
            for year in years:
                self.logger.info(
                    f"Downloading data for station ID: {weather_station} for year: {year}"
                )
                # start and end date need to be in format YYYY-MM-DD
                raw_data = requests.get(f"{base_url}{weather_station}-{year}.zip")
                if raw_data.status_code == 200:
                    # Save to temp file
                    temp_file_path = (
                        Path(self.extension_temp_dir_path)
                        / f"station_{weather_station}.zip"
                    )
                    with open(temp_file_path, "wb") as temp_file:
                        temp_file.write(raw_data.content)

                    # Unzip the file
                    shutil.unpack_archive(
                        temp_file_path,
                        extract_dir=self.extension_temp_dir_path,
                    )

                    # Remove the zip file after extraction
                    temp_file_path.unlink()

    def process_weather_data(self):
        """Process downloaded weather data from Stuttgart stations."""
        self.logger.info("Processing weather data for Stuttgart stations...")

        weather_data = []
        years = range(self.start_datetime.year, self.end_datetime.year + 1)

        for station_name in self.station_ids:
            for year in years:
                # open the existing workbook
                self.logger.info(f"Loading workbook: {station_name} for year {year}")
                workbook = openpyxl.load_workbook(
                    f"./{self.extension_temp_dir_path}/{station_name}_{year}.xlsx"
                )

                if "S-Bad-Cannstatt_Branddirektion_Halbstd.-Werte" in station_name:
                    sheet_names = workbook.sheetnames[:-1]
                elif "S-Mitte_AfU_Halbstd.-Werte" in station_name:
                    sheet_names = workbook.sheetnames[:-1]
                elif "S-Vai-Uni" in station_name:
                    sheet_names = [
                        sheet for sheet in workbook.sheetnames if "(10Min)" in sheet
                    ]
                elif "S-Mue-HKW" in station_name:
                    sheet_names = [
                        sheet for sheet in workbook.sheetnames if "(10Min)" in sheet
                    ]

                self.logger.info(
                    f"Found {len(sheet_names)} sheets in the workbook {station_name} for year {year}."
                )
                for sheet_name in sheet_names:
                    sheet = workbook[sheet_name]
                    data = []
                    self.logger.info(f"Processing sheet: {sheet_name}")
                    # get data from column A
                    time = []
                    for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
                        try:
                            time.append(row[0] if row[0] is not None else None)
                        except Exception:
                            pass

                    # get data from column C
                    temperature = []
                    for row in sheet.iter_rows(
                        min_row=2, min_col=3, max_col=3, values_only=True
                    ):
                        try:
                            temperature.append(
                                float(row[0]) if row[0] is not None else None
                            )
                        except ValueError:
                            continue

                    # get data from column F (or I)
                    if "S-Bad-Cannstatt_Branddirektion_Halbstd.-Werte" in station_name:
                        humidity_start_col = 6
                    elif "S-Mitte_AfU_Halbstd.-Werte" in station_name:
                        humidity_start_col = 6
                    elif "S-Vai-Uni" in station_name:
                        humidity_start_col = 9
                    elif "S-Mue-HKW" in station_name:
                        humidity_start_col = 9

                    humidity = []
                    for row in sheet.iter_rows(
                        min_row=2,
                        min_col=humidity_start_col,
                        max_col=humidity_start_col,
                        values_only=True,
                    ):
                        try:
                            humidity.append(
                                float(row[0]) if row[0] is not None else None
                            )
                        except ValueError:
                            continue

                    # get data from column R or set None
                    if "S-Bad-Cannstatt_Branddirektion_Halbstd.-Werte" in station_name:
                        precipitation_col = -1
                    elif "S-Mitte_AfU_Halbstd.-Werte" in station_name:
                        precipitation_col = -1
                    elif "S-Vai-Uni" in station_name:
                        precipitation_col = 18
                    elif "S-Mue-HKW" in station_name:
                        precipitation_col = 18
                    precipitation = []
                    if precipitation_col != -1:
                        for row in sheet.iter_rows(
                            min_row=2,
                            min_col=precipitation_col,
                            max_col=precipitation_col,
                            values_only=True,
                        ):
                            try:
                                precipitation.append(
                                    float(row[0]) if row[0] is not None else None
                                )
                            except ValueError:
                                continue
                    else:
                        precipitation = [None] * (sheet.max_row - 1)

                    # get data from column O (or S)
                    if "S-Bad-Cannstatt_Branddirektion_Halbstd.-Werte" in station_name:
                        wind_speed_col = 15
                    elif "S-Mitte_AfU_Halbstd.-Werte" in station_name:
                        wind_speed_col = 15
                    elif "S-Vai-Uni" in station_name:
                        wind_speed_col = 19
                    elif "S-Mue-HKW" in station_name:
                        wind_speed_col = 19
                    wind_speed = []
                    for row in sheet.iter_rows(
                        min_row=2,
                        min_col=wind_speed_col,
                        max_col=wind_speed_col,
                        values_only=True,
                    ):
                        try:
                            wind_speed.append(
                                float(row[0]) if row[0] is not None else None
                            )
                        except ValueError:
                            continue

                    # get data from column R (or V)
                    if "S-Bad-Cannstatt_Branddirektion_Halbstd.-Werte" in station_name:
                        wind_direction_col = 18
                    elif "S-Mitte_AfU_Halbstd.-Werte" in station_name:
                        wind_direction_col = 18
                    elif "S-Vai-Uni" in station_name:
                        wind_direction_col = 21
                    elif "S-Mue-HKW" in station_name:
                        wind_direction_col = 21
                    wind_direction = []
                    for row in sheet.iter_rows(
                        min_row=2,
                        min_col=wind_direction_col,
                        max_col=wind_direction_col,
                        values_only=True,
                    ):
                        try:
                            wind_direction.append(
                                int(float(row[0])) if row[0] is not None else None
                            )
                        except ValueError:
                            continue

                    # HEADER DROPS
                    if "S-Bad-Cannstatt_Branddirektion_Halbstd.-Werte" in station_name:
                        time = time[6:]
                        temperature = temperature[4:]
                        humidity = humidity[4:]
                        wind_speed = wind_speed[4:]
                        wind_direction = wind_direction[4:]
                    elif "S-Mitte_AfU_Halbstd.-Werte" in station_name:
                        time = time[6:]
                        temperature = temperature[4:]
                        humidity = humidity[4:]
                        wind_speed = wind_speed[4:]
                        wind_direction = wind_direction[4:]
                        precipitation = precipitation[4:]
                    elif "S-Vai-Uni" in station_name:
                        time = time[3:]
                        temperature = temperature[3:]
                        humidity = humidity[3:]
                        wind_speed = wind_speed[3:]
                        wind_direction = wind_direction[3:]
                    elif "S-Mue-HKW" in station_name:
                        time = time[3:]
                        temperature = temperature[1:]
                        humidity = humidity[1:]
                        wind_speed = wind_speed[1:]
                        wind_direction = wind_direction[1:]
                        precipitation = precipitation[1:]

                    sheet_data = list(
                        zip(
                            time,
                            temperature,
                            humidity,
                            wind_speed,
                            wind_direction,
                            precipitation,
                        )
                    )

                    # LAST ROW DROPS (Summary)
                    if "S-Bad-Cannstatt_Branddirektion_Halbstd.-Werte" in station_name:
                        sheet_data = sheet_data[:-1]
                    elif "S-Mitte_AfU_Halbstd.-Werte" in station_name:
                        sheet_data = sheet_data[:-1]
                    elif "S-Vai-Uni" in station_name:
                        pass
                    elif "S-Mue-HKW" in station_name:
                        pass

                    for row in sheet_data:
                        timestamp = (
                            int(row[0].timestamp()) if row[0] is not None else None
                        )

                        # timestamp is in Europe/Berlin timezone, convert to UTC
                        if timestamp is not None:
                            berlin_tz = pytz.timezone("Europe/Berlin")
                            dt_berlin = datetime.datetime.fromtimestamp(
                                timestamp, berlin_tz
                            )
                            dt_utc = dt_berlin.astimezone(pytz.UTC)
                            timestamp = int(dt_utc.timestamp())

                        # check if timestamp is within start and end timestamp
                        if (
                            timestamp is None
                            or timestamp < self.start_timestamp
                            or timestamp > self.end_timestamp
                        ):
                            continue

                        lat, lon = {
                            "S-Mitte_AfU_Halbstd.-Werte": (48.774, 9.186),
                            "S-Bad-Cannstatt_Branddirektion_Halbstd.-Werte": (
                                48.797,
                                9.219,
                            ),
                            "S-Mue-HKW": (48.847, 9.241),
                            "S-Vai-Uni": (48.747, 9.107),
                        }[station_name]
                        temperature = row[1]
                        humidity = row[2]
                        wind_speed = row[3]
                        wind_direction = row[4]
                        precipitation = row[5]

                        # if all are None, skip
                        if (
                            temperature is None
                            and humidity is None
                            and wind_speed is None
                            and wind_direction is None
                            and precipitation is None
                        ):
                            continue

                        weather_data.append(
                            (
                                timestamp,
                                round(lat, 3),
                                round(lon, 3),
                                temperature,
                                humidity,
                                precipitation,
                                wind_speed,
                                wind_direction,
                            )
                        )

        self.logger.info(f"Processed {len(weather_data)} weather data records")
        self.weather_data = weather_data

    def _get_unique_location_coordinates(self):
        """Extract unique coordinates and append to location_coordinates.parquet"""
        try:
            # Define the meta data file path first
            meta_dir = Path(self.meta_data_dir_path)
            meta_dir.mkdir(parents=True, exist_ok=True)
            coordinates_file = meta_dir / "location_coordinates.parquet"

            # Get unique coordinates from weather data
            unique_coords = set()
            for data_row in self.weather_data:
                lat = data_row[1]
                lng = data_row[2]
                unique_coords.add((lat, lng))

            if not unique_coords:
                self.logger.warning("No coordinates found in weather data")
                return

            # Convert to list for processing
            new_coords = list(unique_coords)

            # Handle existing file
            if coordinates_file.exists():
                # Read existing data
                existing_data = pq.read_table(coordinates_file)
                existing_coords = set(
                    zip(
                        existing_data.column("lat").to_pylist(),
                        existing_data.column("lng").to_pylist(),
                    )
                )
                max_location_id = max(existing_data.column("location_id").to_pylist())

                # Filter out coordinates that already exist
                new_coords = [
                    (lat, lng)
                    for lat, lng in new_coords
                    if (lat, lng) not in existing_coords
                ]

                if not new_coords:
                    self.logger.info("No new coordinates to add")
                    return

                self.logger.info(
                    f"Adding {len(new_coords)} new coordinates to existing file"
                )
            else:
                max_location_id = 0
                self.logger.info(
                    f"Creating new coordinates file with {len(new_coords)} coordinates"
                )

            # Create data for new coordinates only
            if new_coords:
                new_lats, new_lngs = zip(*new_coords)
                new_location_ids = [
                    max_location_id + i + 1 for i in range(len(new_coords))
                ]

                new_data = pa.table(
                    {"location_id": new_location_ids, "lat": new_lats, "lng": new_lngs}
                )

                # Append to existing file or create new one
                if coordinates_file.exists():
                    existing_data = pq.read_table(coordinates_file)
                    combined_data = pa.concat_tables([existing_data, new_data])
                    pq.write_table(
                        combined_data, coordinates_file, compression="BROTLI"
                    )
                else:
                    pq.write_table(new_data, coordinates_file, compression="BROTLI")

                self.logger.info(f"Saved coordinates to {coordinates_file}")

        except Exception as e:
            self.logger.error(f"Error extracting coordinates: {e}")

    def _map_coordinates_to_location_ids(self):
        """Map coordinates in weather data to location IDs from the parquet file"""
        try:
            # Read the location coordinates file
            meta_dir = Path(self.meta_data_dir_path)
            coordinates_file = meta_dir / "location_coordinates.parquet"

            if not coordinates_file.exists():
                self.logger.error(
                    "location_coordinates.parquet not found. Run _get_unique_location_coordinates first."
                )
                return {}

            # Read the parquet file and create a mapping
            location_data = pq.read_table(coordinates_file)
            coord_to_id = {}

            location_ids = location_data.column("location_id").to_pylist()
            lats = location_data.column("lat").to_pylist()
            lngs = location_data.column("lng").to_pylist()

            for location_id, lat, lng in zip(location_ids, lats, lngs):
                coord_to_id[(round(lat, 3), round(lng, 3))] = location_id

            return coord_to_id

        except Exception as e:
            self.logger.error(f"Error mapping coordinates to location IDs: {e}")
            return {}

    def export_weather_data_to_parquet(self):
        """Export weather data to parquet file with proper schema"""
        try:
            # First ensure location coordinates are up to date
            self._get_unique_location_coordinates()

            # Get coordinate to location ID mapping
            coord_to_id = self._map_coordinates_to_location_ids()
            if not coord_to_id:
                self.logger.error("No coordinate mappings available")
                return

            # Prepare data for export
            export_data = []
            for data_row in self.weather_data:
                lat = data_row[1]
                lng = data_row[2]

                # Get location ID from coordinates
                location_id = coord_to_id.get((lat, lng))
                if location_id is None:
                    continue

                # Initialize weather record with defaults
                weather_record = {
                    "location_id": location_id,
                    "timestamp": data_row[0],
                    "temperature": data_row[3],
                    "humidity": data_row[4],
                    "precipitation": data_row[5],
                    "wind_speed": data_row[6],
                    "wind_direction": data_row[7],
                }
                export_data.append(weather_record)

            location_ids = [r["location_id"] for r in export_data]
            timestamps = [r["timestamp"] for r in export_data]
            temperatures = [r["temperature"] for r in export_data]
            humidities = [r["humidity"] for r in export_data]
            precipitations = [r["precipitation"] for r in export_data]
            wind_speeds = [r["wind_speed"] for r in export_data]
            wind_directions = [r["wind_direction"] for r in export_data]

            # Create PyArrow table
            weather_table = pa.table(
                {
                    "location_id": pa.array(location_ids, type=pa.int32()),
                    "timestamp": pa.array(timestamps, type=pa.timestamp("s", tz="UTC")),
                    "temperature": pa.array(temperatures, type=pa.float64()),
                    "humidity": pa.array(humidities, type=pa.float64()),
                    "precipitation": pa.array(precipitations, type=pa.float64()),
                    "wind_speed": pa.array(wind_speeds, type=pa.float64()),
                    "wind_direction": pa.array(wind_directions, type=pa.int32()),
                }
            )

            # delete duplicates if any
            weather_table = pa_ops.drop_duplicates(
                weather_table,
                ["location_id", "timestamp"],
                keep="first",
            )

            # sort weather_table by location_id and timestamp
            weather_table = weather_table.sort_by(
                [("location_id", "ascending"), ("timestamp", "ascending")]
            )

            # Write to parquet file
            output_file = (
                Path(self.extension_data_dir_path) / "stuttgart_weather.parquet"
            )
            # If file exists, delete it first
            if output_file.exists():
                output_file.unlink()

            pq.write_table(weather_table, output_file, compression="BROTLI")

            # delete the tmp folder
            if Path(self.extension_temp_dir_path).exists():
                shutil.rmtree(Path(self.extension_temp_dir_path))

            self.logger.info(
                f"Exported {len(export_data)} weather records to {output_file}"
            )

        except Exception as e:
            self.logger.error(f"Error exporting weather data to parquet: {e}")
